import io

from babel.dates import format_date
from django.conf import settings
from django.db import transaction
from django.db.utils import IntegrityError
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side

from apps.patients.models import Patient
from apps.researches.models import Research


def research_create(*, research_data, user):
    patient_data = research_data.pop('patient')
    requester_id = research_data.pop('requester')['id']

    patient = Patient.objects.create(**patient_data, created_by=user)

    researches = Research.objects.select_for_update()
    with transaction.atomic():
        total_num = settings.RESEARCH_NUM_OFFSET + researches.count() + 1

        try:
            research = Research.objects.create(
                total_num=total_num, patient=patient,
                requester_id=requester_id, **research_data, created_by=user)
        except IntegrityError as exc:
            if 'violates unique constraint "daily_num_and_date"' in str(exc):
                error_data = {
                    'field': ['Данный номер за день уже используется'],
                }
                return error_data, False

    return research, True


def research_patch(*, research_data, user):
    patient_data = research_data.pop('patient')
    patient_id = patient_data.pop('id')
    Patient.objects.filter(
        id=patient_id).update(**patient_data, updated_by=user)

    research_data['requester_id'] = research_data.pop('requester')['id']

    research_id = research_data.pop('id')
    research = Research.objects.filter(id=research_id)

    try:
        research.update(**research_data, updated_by=user)
    except IntegrityError as exc:
        if 'violates unique constraint "daily_num_and_date"' in str(exc):
            error_data = {
                'field': ['Данный номер за день уже используется'],
            }
            return error_data, False

    return research.get(), True


def research_remove(*, research_id, user):
    Research.objects.filter(id=research_id).update(
        deleted=True, updated_by=user)


def research_export_to_xlsx(
        *, research_id, reagents, series, expiration_date, doctor):
    research = Research.objects.get(id=research_id)

    template_path = settings.BASE_DIR / 'src' / 'apps' / 'researches' / 'templates' / 'result_form.xlsx'  # noqa: #501
    wb = load_workbook(template_path)
    ws = wb.active

    # дата результата
    if research.result_date:
        value = research.result_date
        value = format_date(value, '«dd» MMMM Y г.', locale='ru_RU')
        ws['A3'].value = ws['C22'].value = value

    # номер анализа
    ws['B4'].value = research.total_num

    # ФИО
    if research.patient:
        ws['E3'].value = f'{research.patient.last_name} {research.patient.first_name} {research.patient.middle_name}'  # noqa:E501

    # направлен из
    ws['G5'].value = research.requester.name

    # диагноз
    ws['F7'].value = research.get_reason_display()

    # дата взятия
    if research.collect_date:
        value = research.collect_date.strftime('%d.%m.%Y') + ' г.'
        ws['H9'].value = value

    # результат
    ws['E13'].value = research.get_result_display().upper()

    # набор реагентов
    expiration_date = '.'.join(reversed(expiration_date.split('-')))
    reagents_full = f'{reagents} серия {series}, годен до {expiration_date} г.'
    ws['A15'].value = reagents_full

    # выполнил результат
    ws['A18'].value += doctor

    buffer = io.BytesIO()
    wb.save(buffer)
    wb.close()
    buffer.seek(0)
    return buffer


def researches_daily_export(*, date):
    researches = (
        Research.objects
        .filter(collect_date=date)
        .select_related('patient', 'requester')
        .order_by('daily_num')
    )

    template_path = settings.BASE_DIR / 'src' / 'apps' / 'researches' / 'templates' / 'daily_export.xlsx'  # noqa: #501
    wb = load_workbook(template_path)
    ws = wb.active

    ws['A1'].value = date.strftime('%d.%m.%Y')

    side = Side(style='thin')
    border = Border(left=side, top=side, right=side, bottom=side)
    alignment = Alignment(horizontal='center', vertical='center')

    row = 2
    for research in researches:
        row += 1

        ws[f'A{row}'].value = research.total_num
        ws[f'B{row}'].value = research.daily_num
        ws[f'C{row}'].value = research.collect_date.strftime('%d.%m.%Y')
        ws[f'D{row}'].value = research.analys_taken_date.strftime('%d.%m.%Y')
        ws[f'E{row}'].value = research.patient.full_name
        ws[f'F{row}'].value = research.requester.name
        ws[f'G{row}'].value = research.get_reason_display()
        ws[f'H{row}'].value = research.get_result_display()
        ws[f'I{row}'].value = research.result_date.strftime('%d.%m.%Y')

        for letter in 'ABCDEFGHIK':
            if letter not in 'EF':
                ws[f'{letter}{row}'].alignment = alignment
            ws[f'{letter}{row}'].border = border

    buffer = io.BytesIO()
    wb.save(buffer)
    wb.close()
    buffer.seek(0)
    return buffer
